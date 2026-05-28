#!/usr/bin/env python3
"""
中国银行流水分析工具 v3
用法: python boc_analyzer.py <流水文件.xlsx>
输出: <文件名>_统计数据.xlsx  (多 Sheet)
      <文件名>_分析图表.png
"""

import sys, warnings
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.dates as mdates
import matplotlib.font_manager as fm
from matplotlib.gridspec import GridSpec

from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─────────────────────────────── 字体 ────────────────────────────────────────
def _setup_font():
    cands = [f.fname for f in fm.fontManager.ttflist
             if any(k in f.name for k in
                    ["CJK","Noto","SimHei","STHeiti","WenQuanYi",
                     "PingFang","Microsoft YaHei","Source Han"])]
    if cands:
        plt.rcParams["font.family"] = fm.FontProperties(fname=cands[0]).get_name()
    plt.rcParams["axes.unicode_minus"] = False
_setup_font()

# ─────────────────────────────── 颜色 ────────────────────────────────────────
TEAL   = "#1D9E75"; LTEAL  = "#5DCAA5"
CORAL  = "#D85A30"; LCORAL = "#F0997B"
BLUE   = "#378ADD"; LBLUE  = "#85B7EB"
AMBER  = "#BA7517"; LAMBER = "#EF9F27"
PURPLE = "#7F77DD"; PINK   = "#D4537E"
GREEN  = "#639922"; GRAY   = "#888780"
BG     = "#F9F9F7"; GRID   = "#E4E4E0"
CAT_COLORS = [CORAL,BLUE,TEAL,AMBER,PURPLE,PINK,GREEN,GRAY,LTEAL,LCORAL,LBLUE,LAMBER]

# ─────────────────────────────── 分类规则 ────────────────────────────────────
EXPENSE_RULES = [
    ("餐饮外卖", [
        "水西庄","美团","拉扎斯","梅好回忆","南城香","兰州拉面","汉堡","塔斯汀",
        "包子","早阳","学一水果","米粉","偶遇","烧饼","瑞幸","Cotti","Coffee",
        "奈雪","古茗","蜜雪","沪上阿姨","小果食肆","七鲜","荣记","肉夹馍",
        "临榆炸鸡腿","有里有面","东区学生餐厅","伊兰汇","格格小殿","袁记",
        "炸鸡腿","焖饭","餐厅","餐饮","外卖","北京水西庄","甲乙饼","无名缘",
    ]),
    ("学校缴费",  ["北京交通大学","交大体育馆","交大体育","学平险","学习资料"]),
    ("网购电商",  [
        "京东","拼多多","淘宝","网银在线","盒马","闪魔","猫人","格物致品",
        "步冠","乐橙","华商","抱树熊","超值购","杭州致尊","星美","厦门恒安",
        "米占","濮阳县","大洲","云阳县","雏菊","得物","摩享","华为",
    ]),
    ("交通出行",  ["交通一卡通","携程","中铁","铁路","南京地铁","北京地铁","程支付"]),
    ("便利超市",  [
        "便利店","超市","云裳物联","友宝","旭旺","麦芽","东区便利",
        "校园超市","左邻右舍","状元府","好想来","四方超市","校园便利",
    ]),
    ("娱乐充值",  [
        "抖音钻石","汽水音乐","剪映","BUFF","网易","锐协莎","WPS",
        "百度平台","摩享时光","汽水","腾讯",
    ]),
    ("医疗健康",  ["中医药","医院","药","健寿堂","阿里健康"]),
    ("个人转账",  [
        "微信转账","群收款","张家腾","蒋勤","孙孟浩","颜晓华","蔡梓佳",
        "刘殷秀","张万梅","韩阿涛","乐云龙","董雷","姚海锰","王志彬",
        "付自来","詹宏斌","孙素娥","卓之宇","黄美琪","袁付磊",
    ]),
    ("证券/理财", ["中信建投","银证转账","余额宝","理财"]),
    ("短信服务",  ["短信服务"]),
]

INCOME_RULES = [
    ("家人生活费", ["余朝闩","生活费"]),
    ("助研/工资",  ["工资"]),
    ("商品退款",   ["退款"]),
    ("钱包提现",   ["提现"]),
    ("自助存款",   ["存款"]),
    ("利息",       ["结息"]),
    ("红包",       ["红包"]),
]

def cat_expense(row):
    t = str(row["对方账户"]) + str(row["附言"])
    for cat, kws in EXPENSE_RULES:
        if any(k in t for k in kws):
            return cat
    return "其他杂项"

def cat_income(row):
    t  = str(row["对方账户"]) + str(row["附言"])
    tx = str(row["交易名称"])
    for cat, kws in INCOME_RULES:
        if any(k in t or k in tx for k in kws):
            return cat
    return "其他收入"

# ─────────────────────────────── 读取数据 ────────────────────────────────────
def load(path):
    raw  = pd.read_excel(path, sheet_name=0, header=None)
    mask = raw[0].apply(lambda x: isinstance(x, datetime))
    data = raw[mask].copy()
    data.columns = range(data.shape[1])
    df = pd.DataFrame({
        "日期":     pd.to_datetime(data[0]),
        "金额":     pd.to_numeric(data[4], errors="coerce"),
        "余额":     pd.to_numeric(data[5], errors="coerce"),
        "交易名称": data[7].astype(str),
        "附言":     data[10].astype(str),
        "对方账户": data[12].astype(str),
    }).dropna(subset=["金额"])
    df["月份"]  = df["日期"].dt.to_period("M")
    df["星期"]  = df["日期"].dt.day_name()
    inc = df[df["金额"] > 0].copy()
    exp = df[df["金额"] < 0].copy()
    inc["类别"] = inc.apply(cat_income,  axis=1)
    exp["类别"] = exp.apply(cat_expense, axis=1)
    return df, inc, exp

# ─────────────────────────────── 统计计算 ────────────────────────────────────
def compute(df, inc, exp):
    months = sorted(df["月份"].unique())

    # ── 月度明细 ──────────────────────────────────────────────────────────────
    monthly = []
    for m in months:
        i      = inc[inc["月份"] == m]["金额"]
        e      = exp[exp["月份"] == m]["金额"]
        e_abs  = e.abs()
        d      = df[df["月份"] == m]
        bal_end = float(d.sort_values("日期")["余额"].iloc[-1])
        tot_in  = float(i.sum())
        tot_out = float(e_abs.sum())
        monthly.append({
            "月份":            str(m),
            "收入合计":        round(tot_in, 2),
            "支出合计":        round(tot_out, 2),
            "净收支":          round(tot_in - tot_out, 2),
            "储蓄率%":         round((tot_in - tot_out) / tot_in * 100 if tot_in > 0 else 0, 1),
            "期末余额":        round(bal_end, 2),
            "收入笔数":        int(len(i)),
            "支出笔数":        int(len(e)),
            "总交易笔数":      int(len(d)),
            "单笔最大收入":    round(float(i.max())      if len(i)     else 0, 2),
            "单笔平均收入":    round(float(i.mean())     if len(i)     else 0, 2),
            "单笔最大支出":    round(float(e_abs.max())  if len(e_abs) else 0, 2),
            "单笔最小支出":    round(float(e_abs.min())  if len(e_abs) else 0, 2),
            "单笔平均支出":    round(float(e_abs.mean()) if len(e_abs) else 0, 2),
            "单笔中位支出":    round(float(e_abs.median()) if len(e_abs) else 0, 2),
            "支出标准差":      round(float(e_abs.std())  if len(e_abs) else 0, 2),
            "5元以下支出笔数":    int((e_abs < 5).sum()),
            "5~50元支出笔数":     int(((e_abs >= 5)  & (e_abs < 50)).sum()),
            "50~200元支出笔数":   int(((e_abs >= 50) & (e_abs < 200)).sum()),
            "200元以上支出笔数":  int((e_abs >= 200).sum()),
        })
    monthly_df = pd.DataFrame(monthly)

    # ── 整体汇总 ──────────────────────────────────────────────────────────────
    total_in  = float(inc["金额"].sum())
    total_out = abs(float(exp["金额"].sum()))
    net       = total_in - total_out
    e_abs_all = exp["金额"].abs()
    n_months  = len(months)

    summary_rows = [
        ("统计区间",        f"{df['日期'].min().date()} 至 {df['日期'].max().date()}"),
        ("统计月数",        n_months),
        ("总交易笔数",      len(df)),
        ("───收支总计───",  ""),
        ("收入总计",        round(total_in, 2)),
        ("支出总计",        round(total_out, 2)),
        ("净结余",          round(net, 2)),
        ("整体储蓄率%",     round(net / total_in * 100 if total_in > 0 else 0, 1)),
        ("───月均数据───",  ""),
        ("月均收入",        round(total_in / n_months, 2)),
        ("月均支出",        round(total_out / n_months, 2)),
        ("月均净收支",      round(net / n_months, 2)),
        ("───最值月份───",  ""),
        ("收入最高月",      monthly_df.loc[monthly_df["收入合计"].idxmax(), "月份"]),
        ("收入最低月",      monthly_df.loc[monthly_df["收入合计"].idxmin(), "月份"]),
        ("支出最高月",      monthly_df.loc[monthly_df["支出合计"].idxmax(), "月份"]),
        ("支出最低月",      monthly_df.loc[monthly_df["支出合计"].idxmin(), "月份"]),
        ("───单笔统计───",  ""),
        ("单笔最大收入",    round(float(inc["金额"].max()), 2)),
        ("单笔最大支出",    round(float(e_abs_all.max()), 2)),
        ("单笔平均支出",    round(float(e_abs_all.mean()), 2)),
        ("单笔中位支出",    round(float(e_abs_all.median()), 2)),
        ("单笔支出标准差",  round(float(e_abs_all.std()), 2)),
        ("───余额───",      ""),
        ("期间最高余额",    round(float(df["余额"].max()), 2)),
        ("期间最低余额",    round(float(df["余额"].min()), 2)),
        ("最终余额",        round(float(df.sort_values("日期")["余额"].iloc[-1]), 2)),
        ("───支出规模分布───", ""),
        ("小额支出 <5元 笔数",     int((e_abs_all < 5).sum())),
        ("中小额 5~50元 笔数",     int(((e_abs_all >= 5)  & (e_abs_all < 50)).sum())),
        ("中大额 50~200元 笔数",   int(((e_abs_all >= 50) & (e_abs_all < 200)).sum())),
        ("大额 >200元 笔数",       int((e_abs_all >= 200).sum())),
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["统计项", "数值"])

    # ── 支出分类 ──────────────────────────────────────────────────────────────
    cat_exp_df = (exp.groupby("类别")["金额"]
                    .agg(["sum","count","mean","median","std","max","min"])
                    .rename(columns={"sum":"支出合计","count":"笔数","mean":"平均单笔",
                                     "median":"中位单笔","std":"标准差",
                                     "max":"最大单笔","min":"最小单笔"})
                    .abs().round(2)
                    .sort_values("支出合计", ascending=False))
    total_exp = cat_exp_df["支出合计"].sum()
    cat_exp_df["占比%"]    = (cat_exp_df["支出合计"] / total_exp * 100).round(1)
    cat_exp_df["累计占比%"] = cat_exp_df["占比%"].cumsum().round(1)

    # ── 收入来源 ──────────────────────────────────────────────────────────────
    cat_inc_df = (inc.groupby("类别")["金额"]
                    .agg(["sum","count","mean","max"])
                    .rename(columns={"sum":"收入合计","count":"笔数",
                                     "mean":"平均单笔","max":"最大单笔"})
                    .round(2)
                    .sort_values("收入合计", ascending=False))
    cat_inc_df["占比%"] = (cat_inc_df["收入合计"] / cat_inc_df["收入合计"].sum() * 100).round(1)

    # ── 高额/高频商户 Top25 ───────────────────────────────────────────────────
    merchant = (exp.groupby("对方账户")["金额"]
                   .agg(["sum","count","mean","max","min"])
                   .rename(columns={"sum":"累计支出","count":"交易次数",
                                    "mean":"平均单价","max":"最大一笔","min":"最小一笔"})
                   .abs().round(2)
                   .sort_values("累计支出", ascending=False)
                   .head(25))
    merchant["占总支出%"] = (merchant["累计支出"] / total_out * 100).round(1)

    # ── 按星期 ────────────────────────────────────────────────────────────────
    dow_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    dow_label = ["周一","周二","周三","周四","周五","周六","周日"]
    dow_exp_g = (exp.groupby("星期")["金额"]
                    .agg(["sum","count","mean"])
                    .reindex(dow_order))
    dow_exp_g.index = dow_label
    dow_exp_g.columns = ["累计支出","笔数","平均单笔"]
    dow_exp_g = dow_exp_g.abs().round(2)
    dow_exp_g["占总支出%"] = (dow_exp_g["累计支出"] / total_out * 100).round(1)

    # ── 余额日走势 ────────────────────────────────────────────────────────────
    balance = (df.sort_values("日期")
                 .groupby(df["日期"].dt.date)["余额"].last()
                 .reset_index())
    balance.columns = ["日期","余额"]

    # ── 交易明细（完整） ──────────────────────────────────────────────────────
    detail = df[["日期","金额","余额","交易名称","对方账户","附言"]].copy()
    detail["收支"] = detail["金额"].apply(lambda x: "收入" if x > 0 else "支出")
    detail = detail.sort_values("日期", ascending=False).reset_index(drop=True)

    return (summary_df, monthly_df, cat_exp_df, cat_inc_df,
            merchant, dow_exp_g, balance, detail)

# ─────────────────────────────── 写入 Excel ──────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1D9E75")   # 深绿表头
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SEP_FILL  = PatternFill("solid", fgColor="F2F2F2")   # 分隔行
SEP_FONT  = Font(name="Arial", bold=True, color="444444", size=10)
DATA_FONT = Font(name="Arial", size=10)
THIN      = Side(style="thin", color="D0D0D0")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT   = '#,##0.00'
PCT_FMT   = '0.0"%"'
INT_FMT   = '#,##0'

def _write_sheet(ws, df, title=None):
    """把 DataFrame 写入 worksheet，带格式。"""
    start_row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=max(len(df.columns), 2))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font    = Font(name="Arial", bold=True, color="1D9E75", size=12)
        cell.fill    = PatternFill("solid", fgColor="F0FBF6")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        start_row = 3

    # 表头
    df2 = df.reset_index() if df.index.name else df.reset_index(drop=True)
    cols = list(df2.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[start_row].height = 20

    # 数据行
    for ri, row in enumerate(df2.itertuples(index=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val if not pd.isna(val) else "")
            cell.font   = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            # 斑马纹
            if (ri - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9F9F7")
            col_name = cols[ci - 1]
            # 数字格式
            if isinstance(val, float):
                if "%" in col_name:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = NUM_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, (int, np.integer)):
                cell.number_format = INT_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[ri].height = 18

    # 自动列宽
    for ci, col in enumerate(cols, 1):
        col_vals = [str(col)] + [str(v) for v in df2.iloc[:, ci-1] if v is not None]
        max_len  = max(len(s) for s in col_vals)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len * 1.8 + 2, 40)

    # 冻结首行（表头行）
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _write_summary(ws, df):
    """整体汇总：两列，分隔行特殊处理。"""
    ws.cell(row=1, column=1, value="整体汇总统计").font = Font(
        name="Arial", bold=True, color="1D9E75", size=13)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    for ri, row in enumerate(df.itertuples(index=False), 3):
        label, val = row[0], row[1]
        is_sep = str(label).startswith("───")
        for ci, v in enumerate([label, val], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if is_sep:
                cell.font  = SEP_FONT
                cell.fill  = SEP_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font  = DATA_FONT
                cell.alignment = Alignment(vertical="center",
                                           horizontal="right" if ci == 2 and isinstance(val, (int, float)) else "left")
                if (ri % 2) == 0:
                    cell.fill = PatternFill("solid", fgColor="F9F9F7")
                if isinstance(val, float) and ci == 2:
                    cell.number_format = NUM_FMT
            cell.border = BORDER
        ws.row_dimensions[ri].height = 18

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A3"


def export_xlsx(summary_df, monthly_df, cat_exp_df, cat_inc_df,
                merchant, dow_exp_g, balance, detail, out_path):
    sheets = {
        "整体汇总":       (summary_df,  "整体收支汇总统计"),
        "月度收支明细":   (monthly_df,  "各月收支详细数据"),
        "支出分类统计":   (cat_exp_df,  "按消费类别统计"),
        "收入来源统计":   (cat_inc_df,  "按收入来源统计"),
        "商户TOP25":      (merchant,    "累计支出最高商户"),
        "按星期统计":     (dow_exp_g,   "各星期消费统计"),
        "余额走势":       (balance,     "账户余额日变化"),
        "全部交易明细":   (detail,      "原始交易流水"),
    }

    # 先用 pandas 写基础文件
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, (df, _) in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=True)

    # 再用 openpyxl 重写格式
    wb = load_workbook(out_path)
    for sheet_name, (df, title) in sheets.items():
        del wb[sheet_name]  # 删除 pandas 写的
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "整体汇总":
            _write_summary(ws, df)
        else:
            _write_sheet(ws, df, title=title)

    # 调整 sheet 顺序
    order = list(sheets.keys())
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    wb.save(out_path)
    print(f"  Excel -> {out_path}  ({len(sheets)} 个工作表)")

# ─────────────────────────────── 图表 ────────────────────────────────────────
def _style(ax, title="", ylabel="", xlabel="", ygrid=True, xgrid=False):
    ax.set_facecolor(BG)
    if ygrid: ax.grid(axis="y", color=GRID, lw=0.7, zorder=0)
    if xgrid: ax.grid(axis="x", color=GRID, lw=0.7, zorder=0)
    for s in ["top","right"]: ax.spines[s].set_visible(False)
    ax.spines["left"].set_color(GRID); ax.spines["bottom"].set_color(GRID)
    if title:  ax.set_title(title, fontsize=10.5, fontweight="bold", pad=7, color="#222")
    if ylabel: ax.set_ylabel(ylabel, fontsize=8.5, color=GRAY)
    if xlabel: ax.set_xlabel(xlabel, fontsize=8.5, color=GRAY)
    ax.tick_params(colors=GRAY, labelsize=8)

def _yfmt(ax):
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f"¥{v:,.0f}"))
def _xfmt(ax):
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f"¥{v:,.0f}"))

def make_charts(monthly_df, cat_exp_df, cat_inc_df, merchant,
                dow_exp_g, balance, inc, exp, out_path):

    months  = monthly_df["月份"].tolist()
    x       = list(range(len(months)))
    in_v    = monthly_df["收入合计"].tolist()
    out_v   = monthly_df["支出合计"].tolist()
    net_v   = monthly_df["净收支"].tolist()
    tx_cnt  = monthly_df["总交易笔数"].tolist()
    avg_exp = monthly_df["单笔平均支出"].tolist()
    sav     = monthly_df["储蓄率%"].tolist()

    fig = plt.figure(figsize=(20, 28), facecolor="white")
    fig.suptitle("中国银行流水收支分析报告", fontsize=17, fontweight="bold",
                 y=0.993, color="#111")
    gs = GridSpec(5, 3, figure=fig,
                  hspace=0.52, wspace=0.35,
                  top=0.975, bottom=0.03, left=0.06, right=0.97)

    # ── 图1  月度收支柱 + 净收支折线（跨3列） ───────────────────────────────
    ax1 = fig.add_subplot(gs[0, :])
    w = 0.32
    ax1.bar([i-w/2 for i in x], in_v,  w, color=TEAL,  alpha=0.88, label="收入", zorder=3)
    ax1.bar([i+w/2 for i in x], out_v, w, color=CORAL, alpha=0.88, label="支出", zorder=3)
    ax1.plot(x, net_v, color=BLUE, marker="o", ms=5, lw=2, label="净收支", zorder=5)
    ax1.axhline(0, color="#bbb", lw=0.8)
    ax1.set_xticks(x); ax1.set_xticklabels(months, rotation=25, ha="right")
    _style(ax1, title="月度收入 / 支出 / 净收支"); _yfmt(ax1)
    ax1.legend(fontsize=9, loc="upper left", framealpha=0.7)
    mx1 = max(max(in_v), max(out_v)) if in_v else 1
    for i, nv in enumerate(net_v):
        col = TEAL if nv >= 0 else CORAL
        ax1.text(i, nv + (mx1*0.022 if nv>=0 else -mx1*0.055),
                 f"¥{nv:,.0f}", ha="center", fontsize=7.5, color=col, fontweight="bold")

    # ── 图2  余额日走势（跨3列） ─────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[1, :])
    bal_dates = pd.to_datetime(balance["日期"])
    ax2.fill_between(bal_dates, balance["余额"].values, alpha=0.18, color=BLUE)
    ax2.plot(bal_dates, balance["余额"].values, color=BLUE, lw=1.5, zorder=4)
    _style(ax2, title="账户余额日走势", ylabel="余额 (元)"); _yfmt(ax2)
    ax2.xaxis.set_major_locator(mdates.MonthLocator())
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%y-%m"))
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=25, ha="right")

    # ── 图3  支出分类饼 ──────────────────────────────────────────────────────
    ax3 = fig.add_subplot(gs[2, 0])
    sizes  = cat_exp_df["支出合计"].values
    labels = [f"{k}\n¥{v:,.0f}" for k, v in zip(cat_exp_df.index, sizes)]
    wedges, texts, autots = ax3.pie(
        sizes, labels=labels, autopct="%1.1f%%",
        colors=CAT_COLORS[:len(sizes)], startangle=120,
        pctdistance=0.80, wedgeprops=dict(width=0.55, edgecolor="white", lw=1.5))
    for t in texts:  t.set_fontsize(7.5)
    for t in autots: t.set_fontsize(7.5); t.set_color("white"); t.set_fontweight("bold")
    ax3.set_title("支出类别占比", fontsize=10.5, fontweight="bold", color="#222")

    # ── 图4  支出分类横条 ────────────────────────────────────────────────────
    ax4 = fig.add_subplot(gs[2, 1])
    cnames = list(cat_exp_df.index)
    cvals  = list(cat_exp_df["支出合计"].values)
    hbars  = ax4.barh(cnames, cvals, color=CAT_COLORS[:len(cnames)], alpha=0.88)
    mx4 = max(cvals) if cvals else 1
    for bar, val in zip(hbars, cvals):
        ax4.text(bar.get_width() + mx4*0.01,
                 bar.get_y() + bar.get_height()/2,
                 f"¥{val:,.0f}", va="center", fontsize=8, color="#333")
    ax4.set_facecolor(BG); ax4.grid(axis="x", color=GRID, lw=0.7, zorder=0)
    for s in ["top","right"]: ax4.spines[s].set_visible(False)
    ax4.set_title("各类别累计支出", fontsize=10.5, fontweight="bold", color="#222")
    _xfmt(ax4); ax4.set_xlim(right=mx4*1.22)
    ax4.tick_params(colors=GRAY, labelsize=8.5)

    # ── 图5  收入来源柱 ──────────────────────────────────────────────────────
    ax5 = fig.add_subplot(gs[2, 2])
    inc_names  = list(cat_inc_df.index)
    inc_vals   = list(cat_inc_df["收入合计"].values)
    inc_colors = [TEAL,BLUE,AMBER,CORAL,PURPLE,GRAY,GREEN,PINK]
    bars5 = ax5.bar(range(len(inc_names)), inc_vals,
                    color=inc_colors[:len(inc_names)], alpha=0.88, width=0.55)
    ax5.set_xticks(range(len(inc_names)))
    ax5.set_xticklabels(inc_names, rotation=25, ha="right", fontsize=8)
    _style(ax5, title="收入来源构成", ylabel="元"); _yfmt(ax5)
    mx5 = max(inc_vals) if inc_vals else 1
    for bar, val in zip(bars5, inc_vals):
        ax5.text(bar.get_x()+bar.get_width()/2, bar.get_height()+mx5*0.01,
                 f"¥{val:,.0f}", ha="center", fontsize=7.5, color="#333", fontweight="bold")

    # ── 图6  月交易笔数 ──────────────────────────────────────────────────────
    ax6 = fig.add_subplot(gs[3, 0])
    ax6.bar(x, tx_cnt, color=LBLUE, alpha=0.88, zorder=3)
    ax6.set_xticks(x); ax6.set_xticklabels(months, rotation=25, ha="right")
    _style(ax6, title="各月交易笔数", ylabel="笔")
    for i, v in enumerate(tx_cnt):
        ax6.text(i, v+0.5, str(v), ha="center", fontsize=7.5, color=GRAY)

    # ── 图7  按星期支出 ──────────────────────────────────────────────────────
    ax7 = fig.add_subplot(gs[3, 1])
    dow_vals   = dow_exp_g["累计支出"].fillna(0).tolist()
    dow_lbls   = dow_exp_g.index.tolist()
    dow_colors = [CORAL if d in ["周六","周日"] else BLUE for d in dow_lbls]
    ax7.bar(range(7), dow_vals, color=dow_colors, alpha=0.88, zorder=3)
    ax7.set_xticks(range(7)); ax7.set_xticklabels(dow_lbls)
    _style(ax7, title="按星期累计支出（橙=周末）", ylabel="元"); _yfmt(ax7)

    # ── 图8  支出分布直方图 ──────────────────────────────────────────────────
    ax8 = fig.add_subplot(gs[3, 2])
    e_abs  = exp["金额"].abs()
    e_clip = e_abs[e_abs <= 300]
    ax8.hist(e_clip.values, bins=40, color=CORAL, alpha=0.80, edgecolor="white", lw=0.4)
    _style(ax8, title="单笔支出分布（≤300元部分）", xlabel="金额 (元)", ylabel="频次")
    med_v = float(e_abs.median()); avg_v = float(e_abs.mean())
    ax8.axvline(med_v, color=AMBER, lw=1.5, ls="--", label=f"中位 ¥{med_v:.0f}")
    ax8.axvline(avg_v, color=BLUE,  lw=1.5, ls="--", label=f"均值 ¥{avg_v:.0f}")
    ax8.legend(fontsize=8, framealpha=0.7)

    # ── 图9  月单笔平均支出 ──────────────────────────────────────────────────
    ax9 = fig.add_subplot(gs[4, 0])
    ax9.bar(x, avg_exp, color=LAMBER, alpha=0.88, zorder=3)
    ax9.set_xticks(x); ax9.set_xticklabels(months, rotation=25, ha="right")
    _style(ax9, title="各月单笔平均支出", ylabel="元"); _yfmt(ax9)

    # ── 图10  月储蓄率 ───────────────────────────────────────────────────────
    ax10 = fig.add_subplot(gs[4, 1])
    bar_c = [TEAL if v>=0 else CORAL for v in sav]
    ax10.bar(x, sav, color=bar_c, alpha=0.88, zorder=3)
    ax10.axhline(0, color="#bbb", lw=0.8)
    ax10.set_xticks(x); ax10.set_xticklabels(months, rotation=25, ha="right")
    _style(ax10, title="各月储蓄率 %", ylabel="%")
    ax10.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f"{v:.0f}%"))
    for i, v in enumerate(sav):
        col = TEAL if v>=0 else CORAL
        ax10.text(i, v+(0.5 if v>=0 else -1.5),
                  f"{v:.0f}%", ha="center", fontsize=7.5, color=col, fontweight="bold")

    # ── 图11  Top15 商户横条 ─────────────────────────────────────────────────
    ax11 = fig.add_subplot(gs[4, 2])
    top15  = merchant.head(15)
    tnames = [str(n)[:20] for n in top15.index]
    tvals  = list(top15["累计支出"].values)
    tcnts  = list(top15["交易次数"].values)
    h11    = ax11.barh(tnames, tvals, color=BLUE, alpha=0.85)
    mx11   = max(tvals) if tvals else 1
    for bar, val, cnt in zip(h11, tvals, tcnts):
        ax11.text(bar.get_width()+mx11*0.01,
                  bar.get_y()+bar.get_height()/2,
                  f"¥{val:,.0f} ({cnt}次)", va="center", fontsize=7.5, color="#333")
    ax11.set_facecolor(BG); ax11.grid(axis="x", color=GRID, lw=0.7, zorder=0)
    for s in ["top","right"]: ax11.spines[s].set_visible(False)
    ax11.set_title("累计支出 Top15 商户", fontsize=10.5, fontweight="bold", color="#222")
    _xfmt(ax11); ax11.set_xlim(right=mx11*1.35)
    ax11.tick_params(colors=GRAY, labelsize=8)

    total_in  = float(inc["金额"].sum())
    total_out = abs(float(exp["金额"].sum()))
    fig.text(0.5, 0.005,
             f"统计区间 {monthly_df['月份'].iloc[0]} → {monthly_df['月份'].iloc[-1]}   "
             f"总收入 ¥{total_in:,.2f}   总支出 ¥{total_out:,.2f}   "
             f"净结余 ¥{total_in-total_out:,.2f}   共 {len(inc)+len(exp)} 笔交易",
             ha="center", fontsize=9, color=GRAY, style="italic")

    plt.savefig(out_path, dpi=150, bbox_inches="tight", facecolor="white")
    print(f"  图表 -> {out_path}")

# ─────────────────────────────── 主程序 ──────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("用法: python boc_analyzer.py <流水文件.xlsx>")
        sys.exit(1)

    xlsx = sys.argv[1]
    if not Path(xlsx).exists():
        print(f"错误: 找不到文件 {xlsx}"); sys.exit(1)

    stem    = Path(xlsx).stem
    out_dir = Path(xlsx).parent

    print(f"读取: {xlsx}")
    df, inc, exp = load(xlsx)
    print(f"  共 {len(df)} 条交易记录")

    print("计算统计数据...")
    (summary_df, monthly_df, cat_exp_df, cat_inc_df,
     merchant, dow_exp_g, balance, detail) = compute(df, inc, exp)

    xlsx_out = str(out_dir / f"{stem}_统计数据.xlsx")
    png_out  = str(out_dir / f"{stem}_分析图表.png")

    print("导出 Excel...")
    export_xlsx(summary_df, monthly_df, cat_exp_df, cat_inc_df,
                merchant, dow_exp_g, balance, detail, xlsx_out)

    print("生成图表...")
    make_charts(monthly_df, cat_exp_df, cat_inc_df, merchant,
                dow_exp_g, balance, inc, exp, png_out)

    print(f"\n完成！\n  {xlsx_out}\n  {png_out}")

if __name__ == "__main__":
    main()
